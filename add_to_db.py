from openpyxl import load_workbook

def update_invoice_excel(item, weight_pounds, karat, amount_gold):
    # Load the Excel workbook and select the active sheet
    workbook = load_workbook('path/to/invoice.xlsx')
    sheet = workbook.active

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Write data to the next row
    sheet.cell(row=next_row, column=1).value = item
    sheet.cell(row=next_row, column=2).value = weight_pounds
    sheet.cell(row=next_row, column=3).value = karat
    sheet.cell(row=next_row, column=4).value = amount_gold

    # Save the workbook
    workbook.save('path/to/invoice.xlsx')
